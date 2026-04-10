import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import os

# ── LOAD CLEANED FACT TABLE ──────────────────────────────────────────────────
fact_df = pd.read_csv('data/processed/fact_assessments.csv')

# ── SHEET 2: AT-RISK ONLY ────────────────────────────────────────────────────
# at_risk_flag = 1 means score < 60
at_risk_df = fact_df[fact_df['at_risk_flag'] == 1].copy()

# ── SHEET 3: SCHOOL SUMMARY ──────────────────────────────────────────────────
summary_df = fact_df.groupby('school_id').agg(
    avg_score    = ('score',        'mean'),
    pass_rate_pct= ('pass_fail',    lambda x: round((x == 'Pass').sum() / len(x) * 100, 1)),
    at_risk_count= ('at_risk_flag', 'sum'),
    total_records= ('student_id',   'count')
).reset_index()

summary_df['avg_score'] = summary_df['avg_score'].round(1)

# ── WRITE TO EXCEL ───────────────────────────────────────────────────────────
os.makedirs('outputs', exist_ok=True)
output_path = 'outputs/stakeholder_report.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    fact_df.to_excel(   writer, sheet_name='All Students',  index=False)
    at_risk_df.to_excel(writer, sheet_name='At-Risk Only',  index=False)
    summary_df.to_excel(writer, sheet_name='School Summary',index=False)

# ── APPLY FORMATTING ─────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(output_path)

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
bold_font   = Font(bold=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Bold the header row
    for cell in ws[1]:
        cell.font = bold_font

    # On At-Risk Only sheet: highlight every data row yellow
    if sheet_name == 'At-Risk Only':
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = yellow_fill

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

wb.save(output_path)

print("=== EXCEL EXPORT COMPLETE ===")
print(f"Sheet 1 — All Students : {len(fact_df):,} rows")
print(f"Sheet 2 — At-Risk Only : {len(at_risk_df):,} rows (yellow highlighted)")
print(f"Sheet 3 — School Summary: {len(summary_df)} schools")
print(f"\nFile saved: {output_path}")